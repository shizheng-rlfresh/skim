"""File preview routing and non-trajectory preview widgets for skim.

This module owns file reading, generic preview fallbacks, JSON preview dispatch,
and the preview pane container used by the outer app shell. It does not own the
outer multi-pane app behavior or trajectory-specific rendering internals.
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from typing import Any

from rich.syntax import Syntax
from rich.table import Table
from rich.text import Text
from textual.app import ComposeResult
from textual.containers import Vertical, VerticalScroll
from textual.css.query import NoMatches
from textual.widget import Widget
from textual.widgets import Collapsible, Markdown, Static

from ..core.previewing import looks_like_notebook as _core_looks_like_notebook
from ..core.review import FILE_ANNOTATION_KEY, AnnotationRecord, AnnotationStore
from .scrolling import DragScrollMixin
from .trajectory import JsonInspector, TrajectoryViewer

SYNTAX_MAP = {
    ".py": "python",
    ".json": "json",
    ".ipynb": "json",
    ".js": "javascript",
    ".ts": "typescript",
    ".html": "html",
    ".css": "css",
    ".yaml": "yaml",
    ".yml": "yaml",
    ".toml": "toml",
    ".sh": "bash",
    ".bash": "bash",
    ".rs": "rust",
    ".go": "go",
    ".sql": "sql",
    ".xml": "xml",
    ".csv": "csv",
}
MARKDOWN_EXTENSIONS = {".md", ".markdown", ".mdown"}
JSON_EXTENSIONS = {".json"}
NOTEBOOK_EXTENSIONS = {".ipynb"}
XLSX_EXTENSIONS = {".xlsx"}
MAX_FILE_SIZE = 1_000_000
MAX_JSON_FILE_SIZE = 10_000_000
MAX_CSV_ROWS = 20
MAX_CSV_COLS = 8
MAX_CSV_CELL_WIDTH = 24


@dataclass(frozen=True)
class XlsxSheetPreviewData:
    """Display-ready preview metadata for one workbook sheet."""

    name: str
    columns: list[str]
    rows: list[list[str]]
    row_count: int
    column_count: int
    truncated_rows: bool
    truncated_columns: bool
    empty: bool


@dataclass(frozen=True)
class XlsxPreviewData:
    """Display-ready preview metadata for one workbook."""

    name: str
    sheets: list[XlsxSheetPreviewData]


class CsvPreview(Vertical):
    """Dual CSV preview with a readable table and raw text fallback."""

    def __init__(self, content: str) -> None:
        """Initialize the CSV preview from raw file content."""
        super().__init__(classes="csv-preview")
        parsed = _parse_csv(content)
        if isinstance(parsed, str):
            self._widgets = _csv_parse_error_widgets(content, parsed)
        else:
            self._widgets = _csv_preview_widgets(content, parsed)

    def compose(self) -> ComposeResult:
        """Compose the CSV preview widgets."""
        yield from self._widgets


class XlsxPreview(Vertical):
    """Workbook preview with one capped table per sheet."""

    def __init__(self, workbook: XlsxPreviewData) -> None:
        """Initialize the workbook preview from shared parsed workbook data."""
        super().__init__(classes="xlsx-preview")
        self._widgets = _xlsx_preview_widgets(workbook)

    def compose(self) -> ComposeResult:
        """Compose the workbook preview widgets."""
        yield from self._widgets


class FileAnnotationStatus(Static):
    """Compact file-level annotation summary for non-JSON previews."""

    def __init__(
        self,
        annotations: tuple[AnnotationRecord, ...],
        *,
        selected_annotation_id: str | None = None,
        annotation_mode: bool = False,
        **kwargs: Any,
    ) -> None:
        """Initialize the status widget from stored file annotations."""
        super().__init__("", **kwargs)
        self.update(
            _file_annotation_status_text(
                annotations,
                selected_annotation_id=selected_annotation_id,
                annotation_mode=annotation_mode,
            )
        )


class PreviewPane(DragScrollMixin, VerticalScroll, can_focus=True):
    """Scrollable panel that shows file contents."""

    def __init__(self, **kwargs: Any) -> None:
        """Initialize an empty preview pane."""
        super().__init__(**kwargs)
        self._init_drag_scroll()
        self.current_path: Path | None = None
        self.selected_annotation_ids: dict[str, str] = {}
        self.file_annotation_mode = False

    def show_placeholder(self, message: str = "Select a file") -> None:
        """Show placeholder text when no file is selected."""
        self.current_path = None
        self.file_annotation_mode = False
        self.remove_children()
        self.mount(Static(Text(message, style="dim italic")))

    def show_file(self, path: Path) -> None:
        """Render a file into the pane."""
        same_file = self.current_path is not None and self.current_path.resolve() == path.resolve()
        self.current_path = path
        if not same_file:
            self.file_annotation_mode = False
        self.remove_children()
        browse_root = getattr(self.app, "browse_path", path.parent)
        widgets = render_file(path, browse_root=browse_root)
        store = getattr(self.app, "review_store", None)
        if widgets and not isinstance(widgets[0], (JsonInspector, TrajectoryViewer)):
            if isinstance(store, AnnotationStore):
                annotations = store.annotations_for_path(path, FILE_ANNOTATION_KEY)
                widgets = [
                    FileAnnotationStatus(
                        annotations,
                        selected_annotation_id=self.selected_file_annotation_id(annotations),
                        annotation_mode=self.file_annotation_mode,
                        classes="annotation-status-panel",
                    ),
                    *widgets,
                ]
        for widget in widgets:
            self.mount(widget)
        self.scroll_home(animate=False)
        if (
            widgets
            and isinstance(widgets[0], (JsonInspector, TrajectoryViewer))
            and self.id == getattr(self.app, "active_pane_id", None)
        ):
            self.call_after_refresh(widgets[0].focus_tree_mode)

    def on_click(self) -> None:
        """Mark this pane as the active preview when clicked."""
        if self.id is not None and hasattr(self.app, "set_active_pane"):
            self.app.set_active_pane(self.id)

    def scroll_content(self, delta: int) -> None:
        """Scroll specialized inner content when present, else scroll the pane."""
        viewer = self.active_json_navigator()
        if viewer is not None:
            viewer.handle_vertical_key(delta)
        else:
            self.scroll_relative(y=delta, animate=False)

    def active_json_navigator(self) -> JsonInspector | TrajectoryViewer | None:
        """Return the mounted JSON tree/detail navigator, if present."""
        try:
            viewer = self.query(JsonInspector).first()
        except NoMatches:
            try:
                viewer = self.query(TrajectoryViewer).first()
            except NoMatches:
                return None
            return viewer if isinstance(viewer, TrajectoryViewer) else None
        return viewer if isinstance(viewer, JsonInspector) else None

    def file_annotation_selection_key(self) -> str | None:
        """Return the pane-local selection key for the current file annotation target."""
        if self.current_path is None:
            return None
        return f"{self.current_path.resolve().as_posix()}::{FILE_ANNOTATION_KEY}"

    def selected_file_annotation_id(self, annotations: tuple[AnnotationRecord, ...]) -> str | None:
        """Return the selected file annotation id, defaulting to the newest."""
        key = self.file_annotation_selection_key()
        if key is None or not annotations:
            if key is not None:
                self.selected_annotation_ids.pop(key, None)
            return None
        selected_id = self.selected_annotation_ids.get(key)
        if any(annotation.id == selected_id for annotation in annotations):
            return selected_id
        self.selected_annotation_ids[key] = annotations[0].id
        return annotations[0].id

    def set_selected_file_annotation_id(self, annotation_id: str | None) -> None:
        """Set or clear the selected file annotation id for the current file."""
        key = self.file_annotation_selection_key()
        if key is None:
            return
        if annotation_id is None:
            self.selected_annotation_ids.pop(key, None)
        else:
            self.selected_annotation_ids[key] = annotation_id


def render_file(path: Path, *, browse_root: Path | None = None) -> list[Widget]:
    """Return a list of widgets for the given file."""
    if not path.is_file():
        return [Static(Text(f"Not a file: {path.name}", style="red"))]

    suffix = path.suffix.lower()
    size = path.stat().st_size
    max_size = (
        MAX_JSON_FILE_SIZE
        if suffix in JSON_EXTENSIONS or suffix in NOTEBOOK_EXTENSIONS
        else MAX_FILE_SIZE
    )
    if size > max_size:
        return [Static(Text(f"{path.name} is too large ({size:,} bytes)", style="red"))]

    if suffix in XLSX_EXTENSIONS:
        return _xlsx_widgets_for_file(path)

    try:
        content = path.read_text(errors="replace")
    except Exception as error:
        return [Static(Text(f"Could not read {path.name}: {error}", style="red"))]

    if suffix in MARKDOWN_EXTENSIONS:
        return [Markdown(content)]

    if suffix in NOTEBOOK_EXTENSIONS:
        try:
            parsed = json.loads(content)
            if _looks_like_notebook(parsed):
                return _notebook_preview_widgets(parsed)
        except json.JSONDecodeError:
            pass

    if suffix in JSON_EXTENSIONS:
        try:
            parsed = json.loads(content)
            return [
                JsonInspector(
                    parsed,
                    source_path=path,
                    review_root=browse_root or path.parent,
                )
            ]
        except json.JSONDecodeError:
            pass

    if suffix == ".csv":
        return [CsvPreview(content)]

    lexer = SYNTAX_MAP.get(suffix)
    if lexer:
        return [Static(Syntax(content, lexer, line_numbers=True, word_wrap=True))]
    return [Static(Text(content))]


def _xlsx_widgets_for_file(path: Path) -> list[Widget]:
    """Return workbook preview widgets for one `.xlsx` file."""
    try:
        workbook = _load_xlsx_preview(path)
    except Exception as error:
        return [Static(Text(f"Could not open {path.name}: {error}", style="red"))]
    return [XlsxPreview(workbook)]


def _parse_csv(content: str) -> list[list[str]] | str:
    """Parse CSV content into rows, returning an error message on failure."""
    try:
        reader = csv.reader(StringIO(content), strict=True)
        return [list(row) for row in reader]
    except csv.Error as error:
        return str(error)


def _csv_preview_widgets(content: str, rows: list[list[str]]) -> list[Widget]:
    """Build the normal CSV preview widgets."""
    if not rows:
        return [
            Static(Text("Empty CSV file", style="dim italic")),
            _raw_csv_section(content),
        ]

    header = rows[0]
    body = rows[1:]
    table = _csv_table(header, body)
    summary = Static(_csv_summary_text(header, body), classes="csv-summary")
    return [
        summary,
        Static(table, classes="csv-table"),
        _raw_csv_section(content),
    ]


def _xlsx_preview_widgets(workbook: XlsxPreviewData) -> list[Widget]:
    """Build the workbook preview widgets."""
    widgets: list[Widget] = [Static(_xlsx_summary_text(workbook), classes="xlsx-summary")]
    for sheet in workbook.sheets:
        widgets.append(Static(_xlsx_sheet_summary_text(sheet), classes="xlsx-sheet-label"))
        if sheet.empty:
            widgets.append(Static(Text("Empty sheet", style="dim italic")))
            continue
        widgets.append(Static(_xlsx_table(sheet.columns, sheet.rows), classes="xlsx-table"))
    if len(workbook.sheets) == 0:
        widgets.append(Static(Text("Empty workbook", style="dim italic")))
    return widgets


def _csv_parse_error_widgets(content: str, error: str) -> list[Widget]:
    """Build the CSV preview widgets for malformed CSV input."""
    return [
        Static(
            Text(f"CSV parse error: {error}", style="bold red"),
            classes="csv-parse-error",
        ),
        _raw_csv_section(content, collapsed=False),
    ]


def _csv_summary_text(header: list[str], body: list[list[str]]) -> Text:
    """Return a short summary for the CSV preview."""
    text = Text()
    text.append("CSV Preview", style="bold")
    text.append(
        f"  {len(body):,} rows x {len(header):,} columns",
        style="dim",
    )
    if len(body) > MAX_CSV_ROWS or len(header) > MAX_CSV_COLS:
        row_count = min(len(body), MAX_CSV_ROWS)
        column_count = min(len(header), MAX_CSV_COLS)
        text.append(
            f"  showing first {row_count} rows and {column_count} columns",
            style="yellow",
        )
    return text


def _xlsx_summary_text(workbook: XlsxPreviewData) -> Text:
    """Return a short summary for the workbook preview."""
    text = Text()
    text.append("Workbook Preview", style="bold")
    text.append(f"  {len(workbook.sheets):,} sheets", style="dim")
    text.append(f"  {workbook.name}", style="dim")
    return text


def _xlsx_sheet_summary_text(sheet: XlsxSheetPreviewData) -> Text:
    """Return a short summary for one sheet preview."""
    text = Text()
    text.append(sheet.name, style="bold")
    if sheet.empty:
        text.append("  empty", style="dim")
        return text
    text.append(f"  {sheet.row_count:,} rows x {sheet.column_count:,} columns", style="dim")
    notices = []
    if sheet.truncated_rows:
        notices.append(f"showing first {MAX_CSV_ROWS} rows")
    if sheet.truncated_columns:
        notices.append(f"showing first {MAX_CSV_COLS} columns")
    if notices:
        text.append(f"  {'; '.join(notices)}", style="yellow")
    return text


def _csv_table(header: list[str], rows: list[list[str]]) -> Table:
    """Return a capped rich table for a CSV preview."""
    display_header = header[:MAX_CSV_COLS]
    table = Table(expand=True)
    for column in display_header:
        table.add_column(_clip_csv_cell(column), overflow="fold")
    if len(header) > MAX_CSV_COLS:
        table.add_column("…", overflow="fold")

    for row in rows[:MAX_CSV_ROWS]:
        display_row = [
            _clip_csv_cell(row[index] if index < len(row) else "")
            for index in range(len(display_header))
        ]
        if len(header) > MAX_CSV_COLS:
            overflow_values = row[MAX_CSV_COLS:]
            display_row.append(_clip_csv_cell(" | ".join(overflow_values)))
        table.add_row(*display_row)

    return table


def _xlsx_table(header: list[str], rows: list[list[str]]) -> Table:
    """Return a table for one workbook sheet preview."""
    table = Table(expand=True)
    for column in header:
        table.add_column(column or " ", overflow="fold")
    for row in rows:
        table.add_row(*row)
    return table


def _clip_csv_cell(value: str) -> str:
    """Clip one CSV cell to a conservative width for TUI readability."""
    value = " ".join(value.splitlines())
    if len(value) <= MAX_CSV_CELL_WIDTH:
        return value
    return value[: MAX_CSV_CELL_WIDTH - 1].rstrip() + "…"


def _load_xlsx_preview(path: Path) -> XlsxPreviewData:
    """Load one workbook into display-ready preview data."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = [_xlsx_sheet_preview_data(sheet) for sheet in workbook.worksheets]
    finally:
        workbook.close()
    return XlsxPreviewData(name=path.name, sheets=sheets)


def _xlsx_sheet_preview_data(sheet: Any) -> XlsxSheetPreviewData:
    """Return display-ready preview data for one worksheet."""
    sampled_rows: list[list[str]] = []
    sampled_non_empty_rows = 0
    sampled_max_columns = 0
    declared_row_count = max(int(sheet.max_row or 0), 0)
    declared_column_count = max(int(sheet.max_column or 0), 0)
    scan_row_limit = min(declared_row_count, MAX_CSV_ROWS + 1)
    scan_column_limit = min(declared_column_count, MAX_CSV_COLS + 1)

    for row in sheet.iter_rows(
        min_row=1,
        max_row=scan_row_limit,
        min_col=1,
        max_col=scan_column_limit,
        values_only=True,
    ):
        values = [_xlsx_cell_text(value) for value in row]
        while values and values[-1] == "":
            values.pop()
        if not values:
            continue
        sampled_non_empty_rows += 1
        sampled_max_columns = max(sampled_max_columns, len(values))
        if len(sampled_rows) < MAX_CSV_ROWS + 1:
            sampled_rows.append(values)

    row_count = (
        declared_row_count if declared_row_count > MAX_CSV_ROWS + 1 else sampled_non_empty_rows
    )
    column_count = (
        declared_column_count if declared_column_count > MAX_CSV_COLS + 1 else sampled_max_columns
    )

    if row_count == 0 or column_count == 0:
        return XlsxSheetPreviewData(
            name=str(sheet.title),
            columns=[],
            rows=[],
            row_count=0,
            column_count=0,
            truncated_rows=False,
            truncated_columns=False,
            empty=True,
        )

    display_header = [
        _spreadsheet_column_label(index) for index in range(1, min(column_count, MAX_CSV_COLS) + 1)
    ]
    if column_count > MAX_CSV_COLS:
        display_header.append("...")

    display_rows: list[list[str]] = []
    visible_columns = min(column_count, MAX_CSV_COLS)
    for row in sampled_rows[:MAX_CSV_ROWS]:
        padded = row + [""] * max(0, min(scan_column_limit, column_count) - len(row))
        display_row = [_clip_csv_cell(padded[index]) for index in range(visible_columns)]
        if column_count > MAX_CSV_COLS:
            display_row.append(_clip_csv_cell(" | ".join(padded[MAX_CSV_COLS:])))
        display_rows.append(display_row)

    return XlsxSheetPreviewData(
        name=str(sheet.title),
        columns=display_header,
        rows=display_rows,
        row_count=row_count,
        column_count=column_count,
        truncated_rows=row_count > MAX_CSV_ROWS,
        truncated_columns=column_count > MAX_CSV_COLS,
        empty=False,
    )


def _xlsx_cell_text(value: Any) -> str:
    """Normalize one workbook cell value for preview display."""
    if value is None:
        return ""
    return str(value)


def _spreadsheet_column_label(index: int) -> str:
    """Return the spreadsheet column label for one 1-based column index."""
    label = ""
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        label = chr(65 + remainder) + label
    return label


def _raw_csv_section(content: str, *, collapsed: bool = True) -> Collapsible:
    """Return the raw CSV section for a CSV preview."""
    return Collapsible(
        Static(Syntax(content, "csv", line_numbers=True, word_wrap=True), classes="csv-raw"),
        title="Raw CSV",
        collapsed=collapsed,
        classes="csv-raw-section",
    )


def _looks_like_notebook(data: Any) -> bool:
    """Return whether parsed JSON looks like a notebook payload."""
    return _core_looks_like_notebook(data)


def _notebook_preview_widgets(notebook: dict[str, Any]) -> list[Widget]:
    """Build the notebook preview widgets."""
    cells = notebook.get("cells")
    if not isinstance(cells, list):
        cells = []
    language = _notebook_language(notebook.get("metadata"))

    widgets: list[Widget] = [
        Static(_notebook_summary_text(notebook, len(cells), language), classes="notebook-summary")
    ]
    for index, cell in enumerate(cells, start=1):
        widgets.extend(_notebook_cell_widgets(cell, index, language))
    if len(widgets) == 1:
        widgets.append(Static(Text("Empty notebook", style="dim italic")))
    return widgets


def _notebook_language(metadata: Any) -> str:
    """Return the notebook language, defaulting to python."""
    if isinstance(metadata, dict):
        language_info = metadata.get("language_info")
        if isinstance(language_info, dict):
            candidate = language_info.get("name")
            if isinstance(candidate, str) and candidate:
                return candidate
    return "python"


def _notebook_summary_text(notebook: dict[str, Any], cell_count: int, language: str) -> Text:
    """Return a short summary for the notebook preview."""
    text = Text()
    text.append("Notebook Preview", style="bold")
    text.append(f"  {cell_count:,} cells", style="dim")
    nbformat = notebook.get("nbformat")
    nbformat_minor = notebook.get("nbformat_minor")
    if isinstance(nbformat, int):
        version = str(nbformat)
        if isinstance(nbformat_minor, int):
            version += f".{nbformat_minor}"
        text.append(f"  nbformat {version}", style="dim")
    text.append(f"  {language}", style="dim")
    return text


def _notebook_cell_widgets(cell: Any, index: int, language: str) -> list[Widget]:
    """Return flat widgets for one notebook cell and its outputs."""
    if not isinstance(cell, dict):
        return [
            Static(Text(f"Cell {index}", style="bold")),
            Static(Text("Unsupported cell payload", style="yellow")),
        ]
    cell_type = cell.get("cell_type")
    if not isinstance(cell_type, str):
        cell_type = "raw"
    source = _notebook_text(cell.get("source"))

    widgets: list[Widget] = [
        Static(
            Text(f"{cell_type.title()} Cell {index}", style="bold"),
            classes="notebook-cell-label",
        )
    ]
    if cell_type == "markdown":
        widgets.append(Markdown(source or "_Empty markdown cell_"))
    elif cell_type == "code":
        widgets.append(
            Static(
                Syntax(source, language, line_numbers=True, word_wrap=True),
                classes="notebook-code-cell",
            )
        )
    else:
        widgets.append(Static(Text(source), classes="notebook-raw-cell"))

    if cell_type == "code":
        outputs = cell.get("outputs")
        if isinstance(outputs, list):
            for output_index, output in enumerate(outputs, start=1):
                widgets.append(
                    Static(
                        Text(f"Output {index}.{output_index}", style="bold"),
                        classes="notebook-output-label",
                    )
                )
                widgets.append(_notebook_output_widget(output))
    return widgets


def _notebook_output_widget(output: Any) -> Widget:
    """Return a widget for one code-cell output."""
    if not isinstance(output, dict):
        return Static(Text(str(output)))
    if output.get("output_type") == "stream":
        return Static(Text(_notebook_text(output.get("text"))))
    if output.get("output_type") in {"display_data", "execute_result"}:
        data = output.get("data")
        if isinstance(data, dict):
            markdown = data.get("text/markdown")
            if markdown is not None:
                return Markdown(_notebook_text(markdown))
            plain = data.get("text/plain")
            if plain is not None:
                return Static(Text(_notebook_text(plain)))
    traceback = output.get("traceback") if isinstance(output, dict) else None
    if traceback is not None:
        return Static(Text(_notebook_text(traceback), style="red"))
    return Static(Text(json.dumps(output, indent=2)))


def _file_annotation_status_text(
    annotations: tuple[AnnotationRecord, ...],
    *,
    selected_annotation_id: str | None,
    annotation_mode: bool,
) -> Text:
    """Return a compact summary of file-level annotation state."""
    title = Text("File Annotation", style="bold")
    if not annotations:
        title.append("\nNo annotation yet", style="white")
        title.append("\nPress a to annotate this file", style="dim")
        return title
    selected = next(
        (annotation for annotation in annotations if annotation.id == selected_annotation_id),
        annotations[0],
    )
    title.append(
        f"\n{len(annotations)} annotation{'s' if len(annotations) != 1 else ''}",
        style="white",
    )
    for index, annotation in enumerate(annotations, start=1):
        marker = "▶ " if annotation.id == selected.id else "  "
        title.append(
            f"\n{marker}{index}. {annotation.updated_at[:16].replace('T', ' ')} ",
            style="bold green" if annotation.id == selected.id else "dim",
        )
        if annotation.tags:
            title.append(f"[{', '.join(annotation.tags)}] ", style="cyan")
        first_line = (annotation.note or "(empty)").splitlines()[0]
        title.append(first_line[:80] + ("…" if len(first_line) > 80 else ""))
    title.append(
        f"\nTags: {', '.join(selected.tags) if selected.tags else '(none)'}",
        style="cyan",
    )
    title.append(f"\nNote: {selected.note or '(empty)'}", style="white")
    title.append(
        "\nUse ↑↓ to select, Enter to edit, a to add another"
        if annotation_mode
        else "\nPress Enter to browse annotations or a to add another",
        style="dim",
    )
    return title


def _notebook_text(value: Any) -> str:
    """Normalize notebook text content."""
    if isinstance(value, list):
        return "".join(str(item) for item in value)
    if value is None:
        return ""
    return str(value)
