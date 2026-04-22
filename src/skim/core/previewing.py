"""Pure preview classification and file-shape helpers."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from typing import Any

SYNTAX_MAP = {
    ".py": "python",
    ".json": "json",
    ".ipynb": "json",
    ".js": "javascript",
    ".ts": "typescript",
    ".html": "html",
    ".css": "css",
    ".yaml": "yaml",
    ".yml": "yaml",
    ".toml": "toml",
    ".sh": "bash",
    ".bash": "bash",
    ".rs": "rust",
    ".go": "go",
    ".sql": "sql",
    ".xml": "xml",
    ".csv": "csv",
}
MARKDOWN_EXTENSIONS = {".md", ".markdown", ".mdown"}
JSON_EXTENSIONS = {".json"}
NOTEBOOK_EXTENSIONS = {".ipynb"}
XLSX_EXTENSIONS = {".xlsx"}
CSV_EXTENSIONS = {".csv"}
CODE_EXTENSIONS = {
    ".py",
    ".js",
    ".ts",
    ".html",
    ".css",
    ".yaml",
    ".yml",
    ".toml",
    ".sh",
    ".bash",
    ".rs",
    ".go",
    ".sql",
    ".xml",
}
TEXT_EXTENSIONS = {
    "",
    ".txt",
    ".log",
    ".rst",
    ".ini",
    ".cfg",
    ".conf",
}
MAX_FILE_SIZE = 1_000_000
MAX_JSON_FILE_SIZE = 10_000_000
MAX_CSV_ROWS = 20
MAX_CSV_COLS = 8
MAX_CSV_CELL_WIDTH = 24


@dataclass(frozen=True)
class XlsxSheetPreviewData:
    """Display-ready preview metadata for one workbook sheet."""

    name: str
    columns: list[str]
    rows: list[list[str]]
    row_count: int
    column_count: int
    truncated_rows: bool
    truncated_columns: bool
    empty: bool


@dataclass(frozen=True)
class XlsxPreviewData:
    """Display-ready preview metadata for one workbook."""

    name: str
    sheets: list[XlsxSheetPreviewData]


def preview_kind_for_path(file_path: str) -> str:
    """Return the shared preview bucket for one relative file path."""
    suffix = Path(file_path).suffix.lower()
    if suffix in MARKDOWN_EXTENSIONS:
        return "markdown"
    if suffix in JSON_EXTENSIONS:
        return "json"
    if suffix in NOTEBOOK_EXTENSIONS:
        return "notebook"
    if suffix in CSV_EXTENSIONS:
        return "csv"
    if suffix in XLSX_EXTENSIONS:
        return "xlsx"
    if suffix in CODE_EXTENSIONS:
        return "code"
    if suffix in TEXT_EXTENSIONS:
        return "text"
    return "other"


def parse_csv(content: str) -> list[list[str]] | str:
    """Parse CSV content into rows, returning an error message on failure."""
    try:
        reader = csv.reader(StringIO(content), strict=True)
        return [list(row) for row in reader]
    except csv.Error as error:
        return str(error)


def clip_csv_cell(value: str) -> str:
    """Clip one CSV cell to a conservative width for skim previews."""
    value = " ".join(value.splitlines())
    if len(value) <= MAX_CSV_CELL_WIDTH:
        return value
    return value[: MAX_CSV_CELL_WIDTH - 1].rstrip() + "…"


def load_xlsx_preview(path: Path) -> XlsxPreviewData:
    """Load one workbook into display-ready preview data."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = [_xlsx_sheet_preview_data(sheet) for sheet in workbook.worksheets]
    finally:
        workbook.close()
    return XlsxPreviewData(name=path.name, sheets=sheets)


def _xlsx_sheet_preview_data(sheet: Any) -> XlsxSheetPreviewData:
    """Return display-ready preview data for one worksheet."""
    sampled_rows: list[list[str]] = []
    sampled_non_empty_rows = 0
    sampled_max_columns = 0
    declared_row_count = max(int(sheet.max_row or 0), 0)
    declared_column_count = max(int(sheet.max_column or 0), 0)
    scan_row_limit = min(declared_row_count, MAX_CSV_ROWS + 1)
    scan_column_limit = min(declared_column_count, MAX_CSV_COLS + 1)

    for row in sheet.iter_rows(
        min_row=1,
        max_row=scan_row_limit,
        min_col=1,
        max_col=scan_column_limit,
        values_only=True,
    ):
        values = [_xlsx_cell_text(value) for value in row]
        while values and values[-1] == "":
            values.pop()
        if not values:
            continue
        sampled_non_empty_rows += 1
        sampled_max_columns = max(sampled_max_columns, len(values))
        if len(sampled_rows) < MAX_CSV_ROWS + 1:
            sampled_rows.append(values)

    row_count = (
        declared_row_count if declared_row_count > MAX_CSV_ROWS + 1 else sampled_non_empty_rows
    )
    column_count = (
        declared_column_count if declared_column_count > MAX_CSV_COLS + 1 else sampled_max_columns
    )

    if row_count == 0 or column_count == 0:
        return XlsxSheetPreviewData(
            name=str(sheet.title),
            columns=[],
            rows=[],
            row_count=0,
            column_count=0,
            truncated_rows=False,
            truncated_columns=False,
            empty=True,
        )

    display_header = [
        _spreadsheet_column_label(index) for index in range(1, min(column_count, MAX_CSV_COLS) + 1)
    ]
    if column_count > MAX_CSV_COLS:
        display_header.append("...")

    display_rows: list[list[str]] = []
    visible_columns = min(column_count, MAX_CSV_COLS)
    for row in sampled_rows[:MAX_CSV_ROWS]:
        padded = row + [""] * max(0, min(scan_column_limit, column_count) - len(row))
        display_row = [clip_csv_cell(padded[index]) for index in range(visible_columns)]
        if column_count > MAX_CSV_COLS:
            display_row.append(clip_csv_cell(" | ".join(padded[MAX_CSV_COLS:])))
        display_rows.append(display_row)

    return XlsxSheetPreviewData(
        name=str(sheet.title),
        columns=display_header,
        rows=display_rows,
        row_count=row_count,
        column_count=column_count,
        truncated_rows=row_count > MAX_CSV_ROWS,
        truncated_columns=column_count > MAX_CSV_COLS,
        empty=False,
    )


def _xlsx_cell_text(value: Any) -> str:
    """Return a stable string form for one spreadsheet cell."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def _spreadsheet_column_label(index: int) -> str:
    """Return spreadsheet-style A, B, ... AA column labels."""
    label = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        label = chr(65 + remainder) + label
    return label


def looks_like_notebook(data: Any) -> bool:
    """Return whether the parsed JSON looks like a notebook document."""
    if not isinstance(data, dict):
        return False
    if not isinstance(data.get("cells"), list):
        return False
    if not isinstance(data.get("nbformat"), int):
        return False
    metadata = data.get("metadata")
    if metadata is not None and not isinstance(metadata, dict):
        return False
    return all(_looks_like_notebook_cell(cell) for cell in data["cells"])


def _looks_like_notebook_cell(cell: Any) -> bool:
    """Return whether one notebook cell has the expected JSON shape."""
    if not isinstance(cell, dict):
        return False
    metadata = cell.get("metadata")
    if metadata is not None and not isinstance(metadata, dict):
        return False
    source = cell.get("source")
    if source is not None and not _looks_like_notebook_text(source):
        return False
    attachments = cell.get("attachments")
    if attachments is not None and not isinstance(attachments, dict):
        return False
    outputs = cell.get("outputs")
    if outputs is not None and not isinstance(outputs, list):
        return False
    return True


def _looks_like_notebook_text(value: Any) -> bool:
    """Return whether one notebook text payload uses a supported shape."""
    return isinstance(value, str) or (
        isinstance(value, list) and all(isinstance(item, str) for item in value)
    )


def notebook_language(metadata: Any) -> str:
    """Return the notebook language name, defaulting to Python."""
    if isinstance(metadata, dict):
        language_info = metadata.get("language_info")
        if isinstance(language_info, dict):
            name = language_info.get("name")
            if isinstance(name, str) and name:
                return name
        kernelspec = metadata.get("kernelspec")
        if isinstance(kernelspec, dict):
            display_name = kernelspec.get("display_name")
            if isinstance(display_name, str) and display_name:
                lowered = display_name.lower()
                if "python" in lowered:
                    return "python"
    return "python"


def notebook_text(value: Any) -> str:
    """Normalize notebook cell or output text payloads into one string."""
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return "".join(str(item) for item in value)
    return str(value) if value is not None else ""


def loads_if_json(content: str) -> Any | None:
    """Return parsed JSON content, or ``None`` when decoding fails."""
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        return None
